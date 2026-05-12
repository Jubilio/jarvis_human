"""
JARVIS AI OS — Sistema de Agentes com Ferramentas Reais
Powered by LangChain + Ollama (100% local e gratuito)
"""
import os
import subprocess
import json
from datetime import datetime
from pathlib import Path

from langchain_ollama import ChatOllama
from langchain.agents import AgentExecutor, create_react_agent
from langchain_core.prompts import PromptTemplate
from langchain_core.tools import tool
from langchain_community.tools import DuckDuckGoSearchRun
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
from reportlab.lib.colors import HexColor
from reportlab.lib.units import cm

# ==============================================================
# FERRAMENTAS DO AGENTE
# ==============================================================

@tool
def search_internet(query: str) -> str:
    """
    Pesquisa informações na internet usando DuckDuckGo.
    Use quando precisar de dados actuais, notícias ou informações online.
    Input: o termo ou pergunta a pesquisar.
    """
    try:
        search = DuckDuckGoSearchRun()
        result = search.run(query)
        return f"Resultados da pesquisa para '{query}':\n{result}"
    except Exception as e:
        return f"Erro ao pesquisar: {str(e)}"


@tool
def read_pdf(file_path: str) -> str:
    """
    Lê e extrai o texto de um ficheiro PDF.
    Input: o caminho completo para o ficheiro PDF (ex: C:/Users/user/documento.pdf).
    """
    try:
        from pypdf import PdfReader
        path = Path(file_path.strip().strip('"'))
        if not path.exists():
            return f"Ficheiro não encontrado: {file_path}"
        reader = PdfReader(str(path))
        text = ""
        for i, page in enumerate(reader.pages[:15]):  # Máximo 15 páginas
            text += f"\n--- Página {i+1} ---\n{page.extract_text()}"
        return text[:8000]  # Limitar para não sobrecarregar o contexto
    except Exception as e:
        return f"Erro ao ler PDF: {str(e)}"


@tool
def read_excel(file_path: str) -> str:
    """
    Lê e extrai os dados de um ficheiro Excel (.xlsx ou .xls).
    Input: o caminho completo para o ficheiro Excel.
    """
    try:
        import openpyxl
        path = Path(file_path.strip().strip('"'))
        if not path.exists():
            return f"Ficheiro não encontrado: {file_path}"
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        result = []
        for sheet_name in wb.sheetnames[:3]:  # Máximo 3 abas
            ws = wb[sheet_name]
            result.append(f"\n=== Aba: {sheet_name} ===")
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 100:  # Máximo 100 linhas
                    result.append("... (mais dados omitidos)")
                    break
                row_str = " | ".join([str(c) if c is not None else "" for c in row])
                if row_str.strip(" |"):
                    rows.append(row_str)
            result.extend(rows)
        return "\n".join(result)[:8000]
    except Exception as e:
        return f"Erro ao ler Excel: {str(e)}"


@tool
def execute_command(command: str) -> str:
    """
    Executa um comando no sistema operativo (PowerShell no Windows).
    Use APENAS para comandos seguros como: listar ficheiros, ver processos, obter info do sistema.
    NUNCA execute comandos destrutivos (delete, format, rm -rf, etc.).
    Input: o comando a executar (ex: 'dir C:/Users', 'Get-Process', 'ipconfig').
    """
    FORBIDDEN = ['rm ', 'del ', 'format ', 'rmdir', 'rd ', ':(){', 'mkfs', 'shutdown', 'reboot']
    cmd_lower = command.lower()
    for forbidden in FORBIDDEN:
        if forbidden in cmd_lower:
            return f"⛔ COMANDO BLOQUEADO POR SEGURANÇA: '{forbidden}' é uma operação destrutiva."
    
    try:
        result = subprocess.run(
            ['powershell', '-Command', command],
            capture_output=True,
            text=True,
            timeout=15,
            encoding='utf-8',
            errors='ignore'
        )
        output = result.stdout or result.stderr or "Comando executado sem output."
        return output[:4000]
    except subprocess.TimeoutExpired:
        return "Timeout: o comando demorou mais de 15 segundos."
    except Exception as e:
        return f"Erro ao executar comando: {str(e)}"


@tool
def generate_report(content: str) -> str:
    """
    Gera um relatório profissional em PDF com o conteúdo fornecido.
    Use quando o utilizador pedir para gerar, criar ou exportar um relatório.
    Input: o conteúdo completo do relatório em texto (pode incluir secções, dados, etc.).
    Retorna o caminho do ficheiro PDF gerado.
    """
    try:
        reports_dir = Path("./reports")
        reports_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"JARVIS_Relatorio_{timestamp}.pdf"
        filepath = reports_dir / filename
        
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        
        # Estilos customizados
        title_style = ParagraphStyle(
            'JarvisTitle', parent=styles['Title'],
            textColor=HexColor('#00B4D8'),
            fontSize=20, spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'JarvisSubtitle', parent=styles['Normal'],
            textColor=HexColor('#6C757D'),
            fontSize=9, spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'JarvisHeading', parent=styles['Heading2'],
            textColor=HexColor('#0077B6'),
            fontSize=13, spaceBefore=14, spaceAfter=6
        )
        body_style = ParagraphStyle(
            'JarvisBody', parent=styles['Normal'],
            fontSize=10, leading=16, spaceAfter=8
        )
        
        story = []
        
        # Cabeçalho
        story.append(Paragraph("J.A.R.V.I.S AI OS — RELATÓRIO GERADO", title_style))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1, color=HexColor('#00B4D8'), spaceAfter=16))
        
        # Conteúdo (parse básico de secções com #)
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                story.append(Spacer(1, 6))
            elif line.startswith('##'):
                story.append(Paragraph(line.replace('##', '').strip(), heading_style))
            elif line.startswith('#'):
                story.append(Paragraph(line.replace('#', '').strip(), heading_style))
            else:
                # Sanitizar HTML básico
                line = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                story.append(Paragraph(line, body_style))
        
        # Rodapé
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor('#ADB5BD')))
        story.append(Paragraph("Documento gerado automaticamente pelo J.A.R.V.I.S AI OS", subtitle_style))
        
        doc.build(story)
        
        abs_path = str(filepath.resolve())
        return f"✅ Relatório gerado com sucesso!\nFicheiro: {abs_path}"
        
    except Exception as e:
        return f"Erro ao gerar relatório: {str(e)}"


# ==============================================================
# CONSTRUÇÃO DO AGENTE
# ==============================================================

TOOLS = [search_internet, read_pdf, read_excel, execute_command, generate_report]

TOOL_NAMES = ", ".join([t.name for t in TOOLS])
TOOL_DESCRIPTIONS = "\n".join([f"- {t.name}: {t.description.split(chr(10))[0]}" for t in TOOLS])

AGENT_PROMPT = PromptTemplate.from_template("""
Você é o J.A.R.V.I.S., um assistente de IA ultra avançado com acesso a ferramentas reais.
Responda sempre em Português de forma inteligente e com tom futurista. Chame o utilizador de "Mestre".

Você tem acesso às seguintes ferramentas:
{tools}

Use o seguinte formato OBRIGATÓRIO:

Question: a pergunta que precisa responder
Thought: o que preciso fazer para responder
Action: a ferramenta a usar [{tool_names}]
Action Input: o input para a ferramenta
Observation: o resultado da ferramenta
... (repita Thought/Action/Observation quantas vezes precisar)
Thought: Já tenho informação suficiente para responder
Final Answer: a resposta final para o Mestre

Comece!

Question: {input}
Thought: {agent_scratchpad}
""")


def create_jarvis_agent() -> AgentExecutor:
    """Instancia e retorna o agente executor do Jarvis."""
    llm = ChatOllama(model="llama3", temperature=0.3)
    agent = create_react_agent(llm, TOOLS, AGENT_PROMPT)
    executor = AgentExecutor(
        agent=agent,
        tools=TOOLS,
        verbose=True,
        max_iterations=6,
        handle_parsing_errors=True,
        return_intermediate_steps=True
    )
    return executor


def needs_agent(message: str) -> bool:
    """Determina se a mensagem requer o agente com ferramentas ou apenas chat simples."""
    AGENT_KEYWORDS = [
        "pesquisa", "pesquisar", "procura", "procurar", "busca", "buscar",
        "internet", "notícias", "noticias", "online",
        "pdf", "excel", "ficheiro", "arquivo", "documento", "lê ", "ler",
        "executa", "executar", "comando", "abrir", "abre",
        "relatório", "relatorio", "gerar", "exportar", "gera"
    ]
    msg_lower = message.lower()
    return any(kw in msg_lower for kw in AGENT_KEYWORDS)
